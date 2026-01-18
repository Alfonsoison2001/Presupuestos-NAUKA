"""
Modelos de base de datos SQLite para Presupuestos NAUKA
"""
import sqlite3
from pathlib import Path

DATABASE_PATH = Path(__file__).parent / "database.db"

def get_connection():
    """Obtener conexión a la base de datos"""
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_database():
    """Inicializar la base de datos con las tablas necesarias"""
    conn = get_connection()
    cursor = conn.cursor()

    # Tabla de Proyectos
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS proyectos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL UNIQUE,
            descripcion TEXT,
            fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            fecha_modificacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Tabla de Partidas (datos principales del presupuesto)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS partidas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            proyecto_id INTEGER NOT NULL,
            categoria TEXT,
            concepto TEXT,
            detalle TEXT,
            proveedor TEXT,
            unidad TEXT,
            cantidad REAL DEFAULT 0,
            moneda TEXT DEFAULT 'MXN',
            unitario REAL DEFAULT 0,
            importe_sin_iva REAL DEFAULT 0,
            sobrecosto_pct REAL DEFAULT 0,
            sobrecosto_monto REAL DEFAULT 0,
            iva_pct REAL DEFAULT 0,
            iva_monto REAL DEFAULT 0,
            importe_total REAL DEFAULT 0,
            tipo_cambio REAL DEFAULT 1,
            total_mxn REAL DEFAULT 0,
            notas TEXT,
            es_parametro TEXT DEFAULT 'PRESUPUESTO',
            torre TEXT,
            piso TEXT,
            depto TEXT,
            fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            fecha_modificacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (proyecto_id) REFERENCES proyectos(id) ON DELETE CASCADE
        )
    """)

    # Agregar columnas torre, piso, depto si no existen (para bases de datos existentes)
    try:
        cursor.execute("ALTER TABLE partidas ADD COLUMN torre TEXT")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE partidas ADD COLUMN piso TEXT")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE partidas ADD COLUMN depto TEXT")
    except:
        pass

    # Tabla de Glosario - Categorías
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS categorias (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL UNIQUE
        )
    """)

    # Tabla de Glosario - Conceptos
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS conceptos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            categoria_id INTEGER,
            nombre TEXT NOT NULL,
            FOREIGN KEY (categoria_id) REFERENCES categorias(id) ON DELETE CASCADE
        )
    """)

    # Tabla de Tipos de Cambio
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS tipos_cambio (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            moneda TEXT NOT NULL UNIQUE,
            valor REAL NOT NULL DEFAULT 1
        )
    """)

    # Insertar tipos de cambio por defecto
    cursor.execute("""
        INSERT OR IGNORE INTO tipos_cambio (moneda, valor) VALUES
        ('MXN', 1),
        ('USD', 20.5),
        ('EUR', 22)
    """)

    # Crear índices para mejorar rendimiento
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_partidas_proyecto ON partidas(proyecto_id)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_partidas_categoria ON partidas(categoria)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_partidas_concepto ON partidas(concepto)")

    # Tabla de Glosario - Categorías por proyecto
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS glosario_categorias (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            proyecto_id INTEGER NOT NULL,
            nombre TEXT NOT NULL,
            FOREIGN KEY (proyecto_id) REFERENCES proyectos(id) ON DELETE CASCADE,
            UNIQUE(proyecto_id, nombre)
        )
    """)

    # Tabla de Glosario - Conceptos por categoría
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS glosario_conceptos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            categoria_id INTEGER NOT NULL,
            nombre TEXT NOT NULL,
            FOREIGN KEY (categoria_id) REFERENCES glosario_categorias(id) ON DELETE CASCADE,
            UNIQUE(categoria_id, nombre)
        )
    """)

    # Índices para glosario
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_glosario_cat_proyecto ON glosario_categorias(proyecto_id)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_glosario_con_categoria ON glosario_conceptos(categoria_id)")

    # Tabla de Cotizaciones de proveedores
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS cotizaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            proyecto_id INTEGER NOT NULL,
            proveedor TEXT NOT NULL,
            categorias TEXT,
            archivo_nombre TEXT,
            fecha_cotizacion DATE,
            fecha_carga TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            num_items INTEGER DEFAULT 0,
            total REAL DEFAULT 0,
            moneda TEXT DEFAULT 'MXN',
            notas TEXT,
            FOREIGN KEY (proyecto_id) REFERENCES proyectos(id) ON DELETE CASCADE
        )
    """)

    # Tabla de Items/Unitarios extraídos de cotizaciones
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS cotizacion_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cotizacion_id INTEGER NOT NULL,
            codigo TEXT,
            descripcion TEXT NOT NULL,
            unidad TEXT,
            cantidad REAL,
            precio_unitario REAL,
            importe REAL,
            moneda TEXT DEFAULT 'MXN',
            FOREIGN KEY (cotizacion_id) REFERENCES cotizaciones(id) ON DELETE CASCADE
        )
    """)

    # Índices para cotizaciones
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_cot_proveedor ON cotizaciones(proveedor)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_cot_proyecto ON cotizaciones(proyecto_id)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_items_cotizacion ON cotizacion_items(cotizacion_id)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_items_descripcion ON cotizacion_items(descripcion)")

    conn.commit()
    conn.close()
    print("Base de datos inicializada correctamente")

# Funciones CRUD para Proyectos
def crear_proyecto(nombre, descripcion=""):
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO proyectos (nombre, descripcion) VALUES (?, ?)",
            (nombre, descripcion)
        )
        conn.commit()
        return cursor.lastrowid
    except sqlite3.IntegrityError:
        # Si ya existe, obtener su ID
        cursor.execute("SELECT id FROM proyectos WHERE nombre = ?", (nombre,))
        row = cursor.fetchone()
        return row['id'] if row else None
    finally:
        conn.close()

def obtener_proyectos():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM proyectos ORDER BY nombre")
    proyectos = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return proyectos

def obtener_proyecto(proyecto_id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM proyectos WHERE id = ?", (proyecto_id,))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None

def eliminar_proyecto(proyecto_id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM proyectos WHERE id = ?", (proyecto_id,))
    conn.commit()
    conn.close()

# Funciones CRUD para Partidas
def crear_partida(proyecto_id, datos):
    conn = get_connection()
    cursor = conn.cursor()

    # Calcular campos automáticos
    cantidad = float(datos.get('cantidad', 0) or 0)
    unitario = float(datos.get('unitario', 0) or 0)
    importe_sin_iva = cantidad * unitario

    sobrecosto_pct = float(datos.get('sobrecosto_pct', 0) or 0)
    sobrecosto_monto = importe_sin_iva * sobrecosto_pct

    iva_pct = float(datos.get('iva_pct', 0) or 0)
    base_con_sobrecosto = importe_sin_iva + sobrecosto_monto
    iva_monto = base_con_sobrecosto * iva_pct

    importe_total = base_con_sobrecosto + iva_monto

    tipo_cambio = float(datos.get('tipo_cambio', 1) or 1)
    total_mxn = importe_total * tipo_cambio

    cursor.execute("""
        INSERT INTO partidas (
            proyecto_id, categoria, concepto, detalle, proveedor, unidad,
            cantidad, moneda, unitario, importe_sin_iva, sobrecosto_pct,
            sobrecosto_monto, iva_pct, iva_monto, importe_total, tipo_cambio,
            total_mxn, notas, es_parametro, torre, piso, depto
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        proyecto_id,
        datos.get('categoria', ''),
        datos.get('concepto', ''),
        datos.get('detalle', ''),
        datos.get('proveedor', ''),
        datos.get('unidad', ''),
        cantidad,
        datos.get('moneda', 'MXN'),
        unitario,
        importe_sin_iva,
        sobrecosto_pct,
        sobrecosto_monto,
        iva_pct,
        iva_monto,
        importe_total,
        tipo_cambio,
        total_mxn,
        datos.get('notas', ''),
        datos.get('es_parametro', 'PRESUPUESTO'),
        datos.get('torre', ''),
        datos.get('piso', ''),
        datos.get('depto', '')
    ))

    conn.commit()
    partida_id = cursor.lastrowid
    conn.close()
    return partida_id

def obtener_partidas(proyecto_id, categoria=None, concepto=None):
    conn = get_connection()
    cursor = conn.cursor()

    query = "SELECT * FROM partidas WHERE proyecto_id = ?"
    params = [proyecto_id]

    if categoria:
        query += " AND categoria = ?"
        params.append(categoria)

    if concepto:
        query += " AND concepto = ?"
        params.append(concepto)

    query += " ORDER BY categoria, concepto, detalle"

    cursor.execute(query, params)
    partidas = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return partidas

def obtener_partida(partida_id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM partidas WHERE id = ?", (partida_id,))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None

def actualizar_partida(partida_id, datos):
    conn = get_connection()
    cursor = conn.cursor()

    # Calcular campos automáticos
    cantidad = float(datos.get('cantidad', 0) or 0)
    unitario = float(datos.get('unitario', 0) or 0)
    importe_sin_iva = cantidad * unitario

    sobrecosto_pct = float(datos.get('sobrecosto_pct', 0) or 0)
    sobrecosto_monto = importe_sin_iva * sobrecosto_pct

    iva_pct = float(datos.get('iva_pct', 0) or 0)
    base_con_sobrecosto = importe_sin_iva + sobrecosto_monto
    iva_monto = base_con_sobrecosto * iva_pct

    importe_total = base_con_sobrecosto + iva_monto

    tipo_cambio = float(datos.get('tipo_cambio', 1) or 1)
    total_mxn = importe_total * tipo_cambio

    cursor.execute("""
        UPDATE partidas SET
            categoria = ?, concepto = ?, detalle = ?, proveedor = ?, unidad = ?,
            cantidad = ?, moneda = ?, unitario = ?, importe_sin_iva = ?,
            sobrecosto_pct = ?, sobrecosto_monto = ?, iva_pct = ?, iva_monto = ?,
            importe_total = ?, tipo_cambio = ?, total_mxn = ?, notas = ?,
            es_parametro = ?, torre = ?, piso = ?, depto = ?,
            fecha_modificacion = CURRENT_TIMESTAMP
        WHERE id = ?
    """, (
        datos.get('categoria', ''),
        datos.get('concepto', ''),
        datos.get('detalle', ''),
        datos.get('proveedor', ''),
        datos.get('unidad', ''),
        cantidad,
        datos.get('moneda', 'MXN'),
        unitario,
        importe_sin_iva,
        sobrecosto_pct,
        sobrecosto_monto,
        iva_pct,
        iva_monto,
        importe_total,
        tipo_cambio,
        total_mxn,
        datos.get('notas', ''),
        datos.get('es_parametro', 'PRESUPUESTO'),
        datos.get('torre', ''),
        datos.get('piso', ''),
        datos.get('depto', ''),
        partida_id
    ))

    conn.commit()
    conn.close()

def eliminar_partida(partida_id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM partidas WHERE id = ?", (partida_id,))
    conn.commit()
    conn.close()

# Funciones para obtener categorías y conceptos únicos
def obtener_categorias_proyecto(proyecto_id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT categoria FROM partidas
        WHERE proyecto_id = ? AND categoria IS NOT NULL AND categoria != ''
        ORDER BY categoria
    """, (proyecto_id,))
    categorias = [row['categoria'] for row in cursor.fetchall()]
    conn.close()
    return categorias

def obtener_conceptos_proyecto(proyecto_id, categoria=None):
    conn = get_connection()
    cursor = conn.cursor()

    if categoria:
        cursor.execute("""
            SELECT DISTINCT concepto FROM partidas
            WHERE proyecto_id = ? AND categoria = ? AND concepto IS NOT NULL AND concepto != ''
            ORDER BY concepto
        """, (proyecto_id, categoria))
    else:
        cursor.execute("""
            SELECT DISTINCT concepto FROM partidas
            WHERE proyecto_id = ? AND concepto IS NOT NULL AND concepto != ''
            ORDER BY concepto
        """, (proyecto_id,))

    conceptos = [row['concepto'] for row in cursor.fetchall()]
    conn.close()
    return conceptos

# Funciones para resumen/totales
def obtener_resumen_proyecto(proyecto_id):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            categoria,
            COUNT(*) as num_partidas,
            SUM(total_mxn) as total_categoria
        FROM partidas
        WHERE proyecto_id = ?
        GROUP BY categoria
        ORDER BY total_categoria DESC
    """, (proyecto_id,))

    resumen_categorias = [dict(row) for row in cursor.fetchall()]

    cursor.execute("""
        SELECT
            SUM(total_mxn) as total_proyecto,
            COUNT(*) as total_partidas
        FROM partidas
        WHERE proyecto_id = ?
    """, (proyecto_id,))

    totales = dict(cursor.fetchone())

    conn.close()

    return {
        'categorias': resumen_categorias,
        'total_proyecto': totales['total_proyecto'] or 0,
        'total_partidas': totales['total_partidas'] or 0
    }

# Tipos de cambio
def obtener_tipos_cambio():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM tipos_cambio ORDER BY moneda")
    tipos = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return tipos

def actualizar_tipo_cambio(moneda, valor):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT OR REPLACE INTO tipos_cambio (moneda, valor) VALUES (?, ?)",
        (moneda, valor)
    )
    conn.commit()
    conn.close()

# Funciones para resumen agrupado
def obtener_resumen_agrupado(proyecto_id, agrupar_por):
    """
    Obtener resumen agrupado por los campos especificados.
    agrupar_por: lista de campos por los cuales agrupar (ej: ['categoria', 'concepto', 'proveedor', 'torre', 'piso'])
    """
    conn = get_connection()
    cursor = conn.cursor()

    # Campos validos para agrupar
    campos_validos = ['categoria', 'concepto', 'proveedor', 'torre', 'piso', 'depto', 'moneda', 'es_parametro']
    campos_filtrados = [c for c in agrupar_por if c in campos_validos]

    if not campos_filtrados:
        campos_filtrados = ['categoria']

    # Construir la consulta SQL dinamicamente
    campos_select = ', '.join(campos_filtrados)
    campos_group = ', '.join(campos_filtrados)

    query = f"""
        SELECT
            {campos_select},
            COUNT(*) as num_partidas,
            SUM(total_mxn) as total_mxn,
            SUM(importe_sin_iva) as subtotal,
            SUM(iva_monto) as total_iva,
            SUM(sobrecosto_monto) as total_sobrecosto
        FROM partidas
        WHERE proyecto_id = ?
        GROUP BY {campos_group}
        ORDER BY total_mxn DESC
    """

    cursor.execute(query, (proyecto_id,))
    resultados = [dict(row) for row in cursor.fetchall()]

    # Obtener totales generales
    cursor.execute("""
        SELECT
            SUM(total_mxn) as total_proyecto,
            COUNT(*) as total_partidas
        FROM partidas
        WHERE proyecto_id = ?
    """, (proyecto_id,))
    totales = dict(cursor.fetchone())

    conn.close()

    return {
        'agrupado_por': campos_filtrados,
        'resultados': resultados,
        'total_proyecto': totales['total_proyecto'] or 0,
        'total_partidas': totales['total_partidas'] or 0
    }

# Funciones para obtener valores unicos de torre, piso, depto
def obtener_torres_proyecto(proyecto_id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT torre FROM partidas
        WHERE proyecto_id = ? AND torre IS NOT NULL AND torre != ''
        ORDER BY torre
    """, (proyecto_id,))
    torres = [row['torre'] for row in cursor.fetchall()]
    conn.close()
    return torres

def obtener_pisos_proyecto(proyecto_id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT piso FROM partidas
        WHERE proyecto_id = ? AND piso IS NOT NULL AND piso != ''
        ORDER BY piso
    """, (proyecto_id,))
    pisos = [row['piso'] for row in cursor.fetchall()]
    conn.close()
    return pisos

def obtener_proveedores_proyecto(proyecto_id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT proveedor FROM partidas
        WHERE proyecto_id = ? AND proveedor IS NOT NULL AND proveedor != ''
        ORDER BY proveedor
    """, (proyecto_id,))
    proveedores = [row['proveedor'] for row in cursor.fetchall()]
    conn.close()
    return proveedores

# Funciones para resumen jerárquico
def obtener_resumen_jerarquico_nivel1(proyecto_id, filtros=None):
    """
    Obtener nivel 1 del resumen jerárquico: Categorías con totales
    filtros: dict con keys opcionales: torre, piso, depto
    """
    conn = get_connection()
    cursor = conn.cursor()

    # Construir query con filtros opcionales
    query = """
        SELECT
            categoria,
            COUNT(*) as num_partidas,
            SUM(total_mxn) as total_mxn
        FROM partidas
        WHERE proyecto_id = ?
    """
    params = [proyecto_id]

    if filtros:
        if filtros.get('torre'):
            query += " AND torre = ?"
            params.append(filtros['torre'])
        if filtros.get('piso'):
            query += " AND piso = ?"
            params.append(filtros['piso'])
        if filtros.get('depto'):
            query += " AND depto = ?"
            params.append(filtros['depto'])

    query += """
        GROUP BY categoria
        ORDER BY total_mxn DESC
    """

    cursor.execute(query, params)
    categorias = [dict(row) for row in cursor.fetchall()]

    # Obtener totales generales
    query_totales = """
        SELECT
            COUNT(DISTINCT categoria) as num_categorias,
            COUNT(*) as total_partidas,
            SUM(total_mxn) as total_proyecto
        FROM partidas
        WHERE proyecto_id = ?
    """
    params_totales = [proyecto_id]

    if filtros:
        if filtros.get('torre'):
            query_totales += " AND torre = ?"
            params_totales.append(filtros['torre'])
        if filtros.get('piso'):
            query_totales += " AND piso = ?"
            params_totales.append(filtros['piso'])
        if filtros.get('depto'):
            query_totales += " AND depto = ?"
            params_totales.append(filtros['depto'])

    cursor.execute(query_totales, params_totales)
    totales = dict(cursor.fetchone())

    conn.close()

    return {
        'categorias': categorias,
        'num_categorias': totales['num_categorias'] or 0,
        'total_partidas': totales['total_partidas'] or 0,
        'total_proyecto': totales['total_proyecto'] or 0
    }

def obtener_resumen_jerarquico_nivel2(proyecto_id, categoria, filtros=None):
    """
    Obtener nivel 2 del resumen jerárquico: Conceptos de una categoría
    """
    conn = get_connection()
    cursor = conn.cursor()

    query = """
        SELECT
            concepto,
            COUNT(*) as num_partidas,
            SUM(total_mxn) as total_mxn
        FROM partidas
        WHERE proyecto_id = ? AND categoria = ?
    """
    params = [proyecto_id, categoria]

    if filtros:
        if filtros.get('torre'):
            query += " AND torre = ?"
            params.append(filtros['torre'])
        if filtros.get('piso'):
            query += " AND piso = ?"
            params.append(filtros['piso'])
        if filtros.get('depto'):
            query += " AND depto = ?"
            params.append(filtros['depto'])

    query += """
        GROUP BY concepto
        ORDER BY total_mxn DESC
    """

    cursor.execute(query, params)
    conceptos = [dict(row) for row in cursor.fetchall()]
    conn.close()

    return {'conceptos': conceptos}

def obtener_resumen_jerarquico_nivel3(proyecto_id, categoria, concepto, filtros=None):
    """
    Obtener nivel 3 del resumen jerárquico: Detalles de un concepto
    """
    conn = get_connection()
    cursor = conn.cursor()

    query = """
        SELECT
            id,
            detalle,
            proveedor,
            torre,
            piso,
            depto,
            cantidad,
            unidad,
            total_mxn
        FROM partidas
        WHERE proyecto_id = ? AND categoria = ? AND concepto = ?
    """
    params = [proyecto_id, categoria, concepto]

    if filtros:
        if filtros.get('torre'):
            query += " AND torre = ?"
            params.append(filtros['torre'])
        if filtros.get('piso'):
            query += " AND piso = ?"
            params.append(filtros['piso'])
        if filtros.get('depto'):
            query += " AND depto = ?"
            params.append(filtros['depto'])

    query += " ORDER BY total_mxn DESC"

    cursor.execute(query, params)
    detalles = [dict(row) for row in cursor.fetchall()]
    conn.close()

    return {'detalles': detalles}

def obtener_deptos_proyecto(proyecto_id):
    """Obtener departamentos únicos de un proyecto"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT depto FROM partidas
        WHERE proyecto_id = ? AND depto IS NOT NULL AND depto != ''
        ORDER BY depto
    """, (proyecto_id,))
    deptos = [row['depto'] for row in cursor.fetchall()]
    conn.close()
    return deptos

# ============== FUNCIONES CRUD PARA GLOSARIO ==============

def obtener_glosario_proyecto(proyecto_id):
    """Obtener glosario completo de un proyecto con categorías y sus conceptos"""
    conn = get_connection()
    cursor = conn.cursor()

    # Obtener categorías del proyecto
    cursor.execute("""
        SELECT id, nombre FROM glosario_categorias
        WHERE proyecto_id = ?
        ORDER BY nombre
    """, (proyecto_id,))
    categorias = cursor.fetchall()

    resultado = []
    for cat in categorias:
        # Obtener conceptos de cada categoría
        cursor.execute("""
            SELECT id, nombre FROM glosario_conceptos
            WHERE categoria_id = ?
            ORDER BY nombre
        """, (cat['id'],))
        conceptos = [{'id': con['id'], 'nombre': con['nombre']} for con in cursor.fetchall()]

        resultado.append({
            'id': cat['id'],
            'categoria': cat['nombre'],
            'conceptos': conceptos
        })

    conn.close()
    return resultado

def agregar_categoria_glosario(proyecto_id, nombre):
    """Agregar una categoría al glosario del proyecto"""
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO glosario_categorias (proyecto_id, nombre)
            VALUES (?, ?)
        """, (proyecto_id, nombre.strip()))
        conn.commit()
        categoria_id = cursor.lastrowid
        conn.close()
        return {'id': categoria_id, 'nombre': nombre.strip()}
    except sqlite3.IntegrityError:
        conn.close()
        return None  # Ya existe

def eliminar_categoria_glosario(categoria_id):
    """Eliminar una categoría del glosario (cascade elimina conceptos)"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM glosario_categorias WHERE id = ?", (categoria_id,))
    conn.commit()
    conn.close()

def agregar_concepto_glosario(categoria_id, nombre):
    """Agregar un concepto a una categoría del glosario"""
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO glosario_conceptos (categoria_id, nombre)
            VALUES (?, ?)
        """, (categoria_id, nombre.strip()))
        conn.commit()
        concepto_id = cursor.lastrowid
        conn.close()
        return {'id': concepto_id, 'nombre': nombre.strip()}
    except sqlite3.IntegrityError:
        conn.close()
        return None  # Ya existe

def eliminar_concepto_glosario(concepto_id):
    """Eliminar un concepto del glosario"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM glosario_conceptos WHERE id = ?", (concepto_id,))
    conn.commit()
    conn.close()

def importar_glosario_desde_partidas(proyecto_id):
    """Importar categorías y conceptos únicos desde las partidas existentes al glosario"""
    conn = get_connection()
    cursor = conn.cursor()

    # Obtener categorías y conceptos únicos de las partidas
    cursor.execute("""
        SELECT DISTINCT categoria, concepto FROM partidas
        WHERE proyecto_id = ?
        AND categoria IS NOT NULL AND categoria != ''
        AND concepto IS NOT NULL AND concepto != ''
        ORDER BY categoria, concepto
    """, (proyecto_id,))
    partidas_data = cursor.fetchall()

    importados = {'categorias': 0, 'conceptos': 0}

    for row in partidas_data:
        categoria_nombre = row['categoria']
        concepto_nombre = row['concepto']

        # Verificar si la categoría ya existe
        cursor.execute("""
            SELECT id FROM glosario_categorias
            WHERE proyecto_id = ? AND nombre = ?
        """, (proyecto_id, categoria_nombre))
        cat_row = cursor.fetchone()

        if cat_row:
            categoria_id = cat_row['id']
        else:
            # Crear la categoría
            cursor.execute("""
                INSERT INTO glosario_categorias (proyecto_id, nombre)
                VALUES (?, ?)
            """, (proyecto_id, categoria_nombre))
            categoria_id = cursor.lastrowid
            importados['categorias'] += 1

        # Verificar si el concepto ya existe en esa categoría
        cursor.execute("""
            SELECT id FROM glosario_conceptos
            WHERE categoria_id = ? AND nombre = ?
        """, (categoria_id, concepto_nombre))
        con_row = cursor.fetchone()

        if not con_row:
            # Crear el concepto
            cursor.execute("""
                INSERT INTO glosario_conceptos (categoria_id, nombre)
                VALUES (?, ?)
            """, (categoria_id, concepto_nombre))
            importados['conceptos'] += 1

    conn.commit()
    conn.close()
    return importados

def importar_glosario_desde_excel(proyecto_id, archivo_excel):
    """Importar glosario desde la hoja 'Glosario Partidas' de un archivo Excel"""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(archivo_excel, read_only=True, data_only=True)
    except Exception as e:
        return {'error': f'No se pudo abrir el archivo: {str(e)}'}

    # Buscar la hoja de Glosario Partidas primero, luego cualquier hoja con Glosario
    glosario_sheet = None
    for sheet_name in wb.sheetnames:
        if 'glosario' in sheet_name.lower() and 'partida' in sheet_name.lower():
            glosario_sheet = wb[sheet_name]
            break

    if not glosario_sheet:
        for sheet_name in wb.sheetnames:
            if 'glosario' in sheet_name.lower():
                glosario_sheet = wb[sheet_name]
                break

    if not glosario_sheet:
        return {'error': 'No se encontró la hoja de Glosario'}

    conn = get_connection()
    cursor = conn.cursor()

    # Primero eliminar glosario existente del proyecto
    cursor.execute("""
        DELETE FROM glosario_conceptos WHERE categoria_id IN (
            SELECT id FROM glosario_categorias WHERE proyecto_id = ?
        )
    """, (proyecto_id,))
    cursor.execute("DELETE FROM glosario_categorias WHERE proyecto_id = ?", (proyecto_id,))

    importados = {'categorias': 0, 'conceptos': 0}

    # Leer todas las filas
    rows = list(glosario_sheet.iter_rows(values_only=True))

    # Primera pasada: recopilar categorías principales (número entero en B, texto en C)
    categorias_principales = {}  # {nombre_normalizado: nombre_original}

    for row in rows:
        if len(row) < 3:
            continue

        col_b = row[1]
        col_c = row[2]

        if col_b is not None and col_c is not None:
            try:
                num = float(col_b) if not isinstance(col_b, str) else None
                if num is not None and num == int(num) and 1 <= int(num) <= 50:
                    nombre_cat = str(col_c).strip()
                    if nombre_cat and 'categor' not in nombre_cat.lower():
                        nombre_normalizado = nombre_cat.replace('_', ' ').lower().strip()
                        categorias_principales[nombre_normalizado] = nombre_cat
            except (ValueError, TypeError):
                pass

    # Segunda pasada: procesar secciones y conceptos
    categoria_id_actual = None

    for row in rows:
        if len(row) < 3:
            continue

        col_b = row[1]
        col_c = row[2]

        # Detectar encabezado de sección (texto en B que coincide con categoría conocida)
        if col_b is not None and isinstance(col_b, str) and not col_b.startswith('='):
            nombre_seccion = col_b.strip()
            nombre_normalizado = nombre_seccion.replace('_', ' ').lower().strip()

            if nombre_normalizado in categorias_principales:
                # Crear nueva categoría
                nombre_original = categorias_principales[nombre_normalizado]
                cursor.execute("""
                    INSERT INTO glosario_categorias (proyecto_id, nombre)
                    VALUES (?, ?)
                """, (proyecto_id, nombre_original))
                categoria_id_actual = cursor.lastrowid
                importados['categorias'] += 1
                continue

        # Detectar concepto (número decimal en B, texto válido en C)
        if categoria_id_actual and col_b is not None and col_c is not None:
            try:
                num = float(col_b) if not isinstance(col_b, str) else None
                if num is not None:
                    # Verificar que tiene parte decimal significativa
                    parte_decimal = abs(num - int(num))
                    if parte_decimal > 0.001:
                        nombre_concepto = str(col_c).strip()
                        if nombre_concepto and 'categor' not in nombre_concepto.lower():
                            try:
                                cursor.execute("""
                                    INSERT INTO glosario_conceptos (categoria_id, nombre)
                                    VALUES (?, ?)
                                """, (categoria_id_actual, nombre_concepto))
                                importados['conceptos'] += 1
                            except sqlite3.IntegrityError:
                                pass  # Ya existe
            except (ValueError, TypeError):
                pass

    conn.commit()
    conn.close()
    wb.close()

    return importados

# ============== GLOSARIO GLOBAL DE PROVEEDORES ==============

def obtener_proveedores_por_categoria_global():
    """
    Obtener todos los proveedores usados en todos los proyectos,
    agrupados por categoría. Devuelve un diccionario con categorías
    y sus proveedores únicos.
    """
    conn = get_connection()
    cursor = conn.cursor()

    # Obtener todas las combinaciones únicas de categoría-proveedor
    cursor.execute("""
        SELECT DISTINCT categoria, proveedor
        FROM partidas
        WHERE categoria IS NOT NULL AND categoria != ''
        AND proveedor IS NOT NULL AND proveedor != ''
        ORDER BY categoria, proveedor
    """)

    resultados = cursor.fetchall()
    conn.close()

    # Agrupar por categoría
    categorias_dict = {}
    for row in resultados:
        categoria = row['categoria']
        proveedor = row['proveedor']

        if categoria not in categorias_dict:
            categorias_dict[categoria] = []

        if proveedor not in categorias_dict[categoria]:
            categorias_dict[categoria].append(proveedor)

    # Convertir a lista para JSON
    resultado = []
    for categoria in sorted(categorias_dict.keys()):
        resultado.append({
            'categoria': categoria,
            'proveedores': sorted(categorias_dict[categoria]),
            'num_proveedores': len(categorias_dict[categoria])
        })

    return resultado

def obtener_estadisticas_proveedor(proveedor):
    """
    Obtener estadísticas de un proveedor específico:
    - En qué proyectos aparece
    - Total facturado
    - Número de partidas
    """
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            p.nombre as proyecto,
            COUNT(pa.id) as num_partidas,
            SUM(pa.total_mxn) as total_mxn
        FROM partidas pa
        JOIN proyectos p ON pa.proyecto_id = p.id
        WHERE pa.proveedor = ?
        GROUP BY p.id, p.nombre
        ORDER BY total_mxn DESC
    """, (proveedor,))

    proyectos = [dict(row) for row in cursor.fetchall()]

    # Totales generales
    cursor.execute("""
        SELECT
            COUNT(*) as total_partidas,
            SUM(total_mxn) as total_general,
            COUNT(DISTINCT proyecto_id) as num_proyectos
        FROM partidas
        WHERE proveedor = ?
    """, (proveedor,))

    totales = dict(cursor.fetchone())
    conn.close()

    return {
        'proveedor': proveedor,
        'proyectos': proyectos,
        'total_partidas': totales['total_partidas'] or 0,
        'total_general': totales['total_general'] or 0,
        'num_proyectos': totales['num_proyectos'] or 0
    }

# ============== FUNCIONES CRUD PARA COTIZACIONES ==============

import json

def crear_cotizacion(proyecto_id, proveedor, categorias=None, archivo_nombre=None,
                     fecha_cotizacion=None, moneda='MXN', notas=None):
    """Crear una nueva cotización"""
    conn = get_connection()
    cursor = conn.cursor()

    # Convertir lista de categorías a JSON
    categorias_json = json.dumps(categorias) if categorias else None

    cursor.execute("""
        INSERT INTO cotizaciones (proyecto_id, proveedor, categorias, archivo_nombre,
                                  fecha_cotizacion, moneda, notas)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (proyecto_id, proveedor, categorias_json, archivo_nombre,
          fecha_cotizacion, moneda, notas))

    conn.commit()
    cotizacion_id = cursor.lastrowid
    conn.close()
    return cotizacion_id

def obtener_cotizaciones(proyecto_id=None, proveedor=None, categoria=None):
    """Obtener cotizaciones con filtros opcionales"""
    conn = get_connection()
    cursor = conn.cursor()

    query = """
        SELECT c.*, p.nombre as proyecto_nombre
        FROM cotizaciones c
        JOIN proyectos p ON c.proyecto_id = p.id
        WHERE 1=1
    """
    params = []

    if proyecto_id:
        query += " AND c.proyecto_id = ?"
        params.append(proyecto_id)

    if proveedor:
        query += " AND c.proveedor = ?"
        params.append(proveedor)

    if categoria:
        # Buscar en el JSON de categorías
        query += " AND c.categorias LIKE ?"
        params.append(f'%"{categoria}"%')

    query += " ORDER BY c.fecha_carga DESC"

    cursor.execute(query, params)
    cotizaciones = []
    for row in cursor.fetchall():
        cot = dict(row)
        # Parsear categorías de JSON
        if cot.get('categorias'):
            try:
                cot['categorias'] = json.loads(cot['categorias'])
            except:
                cot['categorias'] = []
        else:
            cot['categorias'] = []
        cotizaciones.append(cot)

    conn.close()
    return cotizaciones

def obtener_cotizacion(cotizacion_id):
    """Obtener una cotización por ID"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT c.*, p.nombre as proyecto_nombre
        FROM cotizaciones c
        JOIN proyectos p ON c.proyecto_id = p.id
        WHERE c.id = ?
    """, (cotizacion_id,))
    row = cursor.fetchone()
    conn.close()

    if row:
        cot = dict(row)
        if cot.get('categorias'):
            try:
                cot['categorias'] = json.loads(cot['categorias'])
            except:
                cot['categorias'] = []
        else:
            cot['categorias'] = []
        return cot
    return None

def actualizar_cotizacion(cotizacion_id, datos):
    """Actualizar una cotización"""
    conn = get_connection()
    cursor = conn.cursor()

    categorias_json = json.dumps(datos.get('categorias')) if datos.get('categorias') else None

    cursor.execute("""
        UPDATE cotizaciones SET
            proveedor = ?,
            categorias = ?,
            fecha_cotizacion = ?,
            moneda = ?,
            notas = ?
        WHERE id = ?
    """, (
        datos.get('proveedor'),
        categorias_json,
        datos.get('fecha_cotizacion'),
        datos.get('moneda', 'MXN'),
        datos.get('notas'),
        cotizacion_id
    ))

    conn.commit()
    conn.close()

def eliminar_cotizacion(cotizacion_id):
    """Eliminar una cotización y sus items"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM cotizaciones WHERE id = ?", (cotizacion_id,))
    conn.commit()
    conn.close()

def actualizar_totales_cotizacion(cotizacion_id):
    """Actualizar num_items y total de una cotización"""
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT COUNT(*) as num, COALESCE(SUM(importe), 0) as total
        FROM cotizacion_items WHERE cotizacion_id = ?
    """, (cotizacion_id,))
    row = cursor.fetchone()

    cursor.execute("""
        UPDATE cotizaciones SET num_items = ?, total = ? WHERE id = ?
    """, (row['num'], row['total'], cotizacion_id))

    conn.commit()
    conn.close()

# ============== FUNCIONES CRUD PARA ITEMS DE COTIZACION ==============

def crear_items_cotizacion(cotizacion_id, items):
    """Crear múltiples items para una cotización"""
    conn = get_connection()
    cursor = conn.cursor()

    for item in items:
        cursor.execute("""
            INSERT INTO cotizacion_items (cotizacion_id, codigo, descripcion, unidad,
                                          cantidad, precio_unitario, importe, moneda)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            cotizacion_id,
            item.get('codigo'),
            item.get('descripcion', ''),
            item.get('unidad'),
            item.get('cantidad'),
            item.get('precio_unitario'),
            item.get('importe'),
            item.get('moneda', 'MXN')
        ))

    conn.commit()
    conn.close()

    # Actualizar totales
    actualizar_totales_cotizacion(cotizacion_id)

def obtener_items_cotizacion(cotizacion_id):
    """Obtener items de una cotización"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT * FROM cotizacion_items
        WHERE cotizacion_id = ?
        ORDER BY id
    """, (cotizacion_id,))
    items = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return items

def actualizar_item_cotizacion(item_id, datos):
    """Actualizar un item de cotización"""
    conn = get_connection()
    cursor = conn.cursor()

    # Obtener cotizacion_id antes de actualizar
    cursor.execute("SELECT cotizacion_id FROM cotizacion_items WHERE id = ?", (item_id,))
    row = cursor.fetchone()
    cotizacion_id = row['cotizacion_id'] if row else None

    cursor.execute("""
        UPDATE cotizacion_items SET
            codigo = ?,
            descripcion = ?,
            unidad = ?,
            cantidad = ?,
            precio_unitario = ?,
            importe = ?,
            moneda = ?
        WHERE id = ?
    """, (
        datos.get('codigo'),
        datos.get('descripcion'),
        datos.get('unidad'),
        datos.get('cantidad'),
        datos.get('precio_unitario'),
        datos.get('importe'),
        datos.get('moneda', 'MXN'),
        item_id
    ))

    conn.commit()
    conn.close()

    # Actualizar totales de la cotización
    if cotizacion_id:
        actualizar_totales_cotizacion(cotizacion_id)

def eliminar_item_cotizacion(item_id):
    """Eliminar un item de cotización"""
    conn = get_connection()
    cursor = conn.cursor()

    # Obtener cotizacion_id antes de eliminar
    cursor.execute("SELECT cotizacion_id FROM cotizacion_items WHERE id = ?", (item_id,))
    row = cursor.fetchone()
    cotizacion_id = row['cotizacion_id'] if row else None

    cursor.execute("DELETE FROM cotizacion_items WHERE id = ?", (item_id,))
    conn.commit()
    conn.close()

    # Actualizar totales
    if cotizacion_id:
        actualizar_totales_cotizacion(cotizacion_id)

# ============== FUNCIONES PARA COMPARACION DE UNITARIOS ==============

def obtener_proveedores_cotizaciones():
    """Obtener lista de proveedores únicos que tienen cotizaciones"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT proveedor FROM cotizaciones
        WHERE proveedor IS NOT NULL AND proveedor != ''
        ORDER BY proveedor
    """)
    proveedores = [row['proveedor'] for row in cursor.fetchall()]
    conn.close()
    return proveedores

def comparar_unitarios(proveedor, categoria=None):
    """
    Comparar precios unitarios de un proveedor entre diferentes proyectos.
    Retorna una tabla con descripciones y precios por proyecto.
    """
    conn = get_connection()
    cursor = conn.cursor()

    # Obtener todas las cotizaciones del proveedor
    query = """
        SELECT c.id, c.proyecto_id, p.nombre as proyecto_nombre, c.categorias
        FROM cotizaciones c
        JOIN proyectos p ON c.proyecto_id = p.id
        WHERE c.proveedor = ?
    """
    params = [proveedor]

    if categoria:
        query += " AND c.categorias LIKE ?"
        params.append(f'%"{categoria}"%')

    cursor.execute(query, params)
    cotizaciones = cursor.fetchall()

    if not cotizaciones:
        conn.close()
        return {'items': [], 'proyectos': []}

    # Obtener IDs de cotizaciones
    cotizacion_ids = [c['id'] for c in cotizaciones]
    proyectos_dict = {c['proyecto_id']: c['proyecto_nombre'] for c in cotizaciones}

    # Obtener todos los items de esas cotizaciones
    placeholders = ','.join('?' * len(cotizacion_ids))
    cursor.execute(f"""
        SELECT ci.*, c.proyecto_id
        FROM cotizacion_items ci
        JOIN cotizaciones c ON ci.cotizacion_id = c.id
        WHERE ci.cotizacion_id IN ({placeholders})
        ORDER BY ci.descripcion
    """, cotizacion_ids)

    items = cursor.fetchall()
    conn.close()

    # Agrupar por descripción normalizada
    comparacion = {}
    for item in items:
        desc = item['descripcion'].strip().upper() if item['descripcion'] else ''
        if not desc:
            continue

        if desc not in comparacion:
            comparacion[desc] = {
                'descripcion': item['descripcion'],
                'unidad': item['unidad'],
                'precios': {}
            }

        proyecto_id = item['proyecto_id']
        proyecto_nombre = proyectos_dict.get(proyecto_id, f'Proyecto {proyecto_id}')

        # Guardar precio unitario por proyecto
        if proyecto_nombre not in comparacion[desc]['precios']:
            comparacion[desc]['precios'][proyecto_nombre] = item['precio_unitario']

    # Convertir a lista y calcular diferencias
    resultado = []
    proyectos_lista = sorted(proyectos_dict.values())

    for desc, data in comparacion.items():
        precios = data['precios']
        valores = [v for v in precios.values() if v is not None and v > 0]

        item_resultado = {
            'descripcion': data['descripcion'],
            'unidad': data['unidad'],
            'precios': precios
        }

        if len(valores) >= 2:
            min_precio = min(valores)
            max_precio = max(valores)
            if min_precio > 0:
                diferencia_pct = ((max_precio - min_precio) / min_precio) * 100
                item_resultado['diferencia_pct'] = round(diferencia_pct, 1)
                item_resultado['min_precio'] = min_precio
                item_resultado['max_precio'] = max_precio

        resultado.append(item_resultado)

    # Ordenar por mayor diferencia
    resultado.sort(key=lambda x: x.get('diferencia_pct', 0), reverse=True)

    return {
        'items': resultado,
        'proyectos': proyectos_lista,
        'proveedor': proveedor
    }

if __name__ == "__main__":
    init_database()
